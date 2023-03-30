from openpyxl import load_workbook
import PySimpleGUI as sg
import os

# Define o tema da interface
sg.theme('DarkAmber')

def abrir_script():
    script_path = os.path.join(os.getcwd(),"empresa","desenvolvimento", "TabelaCadastro", "src", "interface.pyw")
    os.system(f'start "" "{script_path}"')

def carregar_dados():
    file_path = os.path.join(os.getcwd(),"empresa","desenvolvimento", "TabelaCadastro", "data", "Tabela.xlsx")
    workbook = load_workbook(file_path)
    worksheet = workbook.active
    data = []
    for row in worksheet.iter_rows(values_only=True):
        data.append(row)
    return data

def criar_tabela(nome_tabela):
    file_path = os.path.join(os.getcwd(),"empresa","analise", "Tabelas", nome_tabela + ".xlsx")
    if not os.path.exists(file_path):
        workbook = load_workbook()
        workbook.save(file_path)

def listar_tabelas():
    tabelas = []
    folder_path = os.path.join(os.getcwd(), "empresa", "analise", "Tabelas")
    for file in os.listdir(folder_path):
        if file.endswith('.xlsx'):
            tabelas.append(file)
    return tabelas

def abrir_tabela(nome_tabela):
    tabela_path = os.path.join(os.getcwd(), "empresa", "analise", "Tabelas", nome_tabela)
    os.system(f'start "" "{tabela_path}"')

def criar_projeto():
    projeto_path = os.path.join(os.getcwd(), "empresa", "desenvolvimento", "TabelaCadastro", "src", "meu_projeto.pyw")
    with open(projeto_path, 'w') as file:
        file.write('print("Meu projeto")')

def listar_projetos():
    projetos = []
    folder_path = os.path.join(os.getcwd(), "empresa", "desenvolvimento", "TabelaCadastro", "src")
    for file in os.listdir(folder_path):
        if file.endswith('.pyw'):
            projetos.append(file)
    return projetos

# Cria o layout da janela
layout = [[sg.Text('Tabela de Cadastro')],
          [sg.Table(values=carregar_dados(), headings=["ID", "Barcode", "Quantidade"], max_col_width=25,
                    auto_size_columns=False, display_row_numbers=False, justification='center', num_rows=20)],
          [sg.Text('Crie uma nova tabela')],
          [sg.Input(key='-NOME_TABELA-'), sg.Button('Criar Tabela')],
          [sg.Listbox(values=listar_tabelas(), size=(30, 10), key='-TABELAS-', enable_events=True)],
          [sg.Text('Crie um novo projeto')],
          [sg.Button('Criar Projeto'), sg.Listbox(values=listar_projetos(), size=(30, 10), key='-PROJETOS-', enable_events=True)],
          [sg.Button('Abrir Projeto', disabled=True, key='-ABRIR-')],
          [sg.Button('Abrir Script')],
          [sg.Button('Fechar')]]

# Cria a janela
window = sg.Window('Tabela de Cadastro', layout)

# Loop de eventos para interação com a janela
while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == 'Fechar':
        break
    elif event == 'Abrir Script':
        abrir_script()
    elif event == 'Criar Tabela':
        nome_tabela = values['-NOME_TABELA-']
        criar_tabela(nome_tabela)
        window['-TABELAS-'].update(values=listar_tabelas())
    elif event == '-TABELAS-':
        tabela_selecionada = values['-TABELAS-'][0]
    if tabela_selecionada:
        window['-ABRIR-'].update(disabled=False)
    elif event == 'Abrir Projeto':
        projeto_selecionado = values['-PROJETOS-'][0]
    if projeto_selecionado:
        abrir_projeto(projeto_selecionado)
    elif event == 'Criar Projeto':
        criar_projeto()
        window['-PROJETOS-'].update(values=listar_projetos())

    window.close()
